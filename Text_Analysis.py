#!/usr/bin/env python
# coding: utf-8

# # Data Extraction and NLP

# Test Assignment

#import the necessary libraries
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
import requests
import os
import re
import string
from nltk.tokenize import word_tokenize, sent_tokenize

#function to extract text and title from url
def extract_text(url):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36'}
    page = requests.get(url, headers=headers)
    soup = BeautifulSoup(page.content, 'html.parser')
    
    content = soup.find_all(attrs={'class': 'td-post-content'})
    content_text = content[0].get_text(strip=True) if content else ''
    
    title = soup.find_all(attrs={'class': 'entry-title'})
    title_text = title[0].get_text(strip=True) if title else ''
    
    text = title_text + '. ' + content_text
    return text

# Function to load positive and negative words from files
def load_words(file_path):
    words = []
    with open(file_path, 'r') as file:
        for line in file:
            words.append(line.strip().lower())
    return words

# Load positive and negative words from your desktop
positive_words = load_words("D:\\projects\\NLP_TextAnalysis\\Assignment\\MasterDictionary\\positive-words.txt")
negative_words = load_words("D:\\projects\\NLP_TextAnalysis\\Assignment\\MasterDictionary\\negative-words.txt")

# Load custom stopword files from your desktop
custom_stopwords_files = [
    "D:\\projects\\NLP_TextAnalysis\\Assignment\\StopWords\\StopWords_Auditor.txt", 
    "D:\\projects\\NLP_TextAnalysis\\Assignment\\StopWords\\StopWords_Currencies.txt",
    "D:\\projects\\NLP_TextAnalysis\\Assignment\\StopWords\\StopWords_DatesandNumbers.txt",
    "D:\\projects\\NLP_TextAnalysis\\Assignment\\StopWords\\StopWords_Generic.txt",
    "D:\\projects\\NLP_TextAnalysis\\Assignment\\StopWords\\StopWords_GenericLong.txt",
    "D:\\projects\\NLP_TextAnalysis\\Assignment\\StopWords\\StopWords_Geographic.txt",
    "D:\\projects\\NLP_TextAnalysis\\Assignment\\StopWords\\StopWords_Names.txt"
]

#function to load stop words
def load_custom_stopwords(files):
    stopwords_set = set()
    for file in files:
        with open(file, 'r') as f:
            stopwords_set.update(f.read().splitlines())
    return stopwords_set
custom_stop_words = load_custom_stopwords(custom_stopwords_files)

#function to clean and tokenize text
def clean_and_tokenize(text):
    words = word_tokenize(text)
    table = str.maketrans('', '', string.punctuation)
    words = [word.translate(table).lower() for word in words]
    words = [word for word in words if word not in custom_stop_words]
    return words

#function to compute variables
def compute_variables(text):
    tokenize_text = clean_and_tokenize(text)
    if not tokenize_text:
        return 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
    
    # positive score
    positive_score = sum(1 for word in tokenize_text if word.lower() in positive_words)
    
    # negative score
    negative_score = sum(1 for word in tokenize_text if word.lower() in negative_words)
    
    # polarity score
    polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)
    
    # subjectivity score
    subjectivity_score = (positive_score - negative_score) / (len(tokenize_text) + 0.000001)
    
    # average sentence length
    sentences = sent_tokenize(text)
    avg_length = sum(len(sent.split()) for sent in sentences) / len(sentences)
    
    # percentage of complex words
    count = 0
    complex_word_count = 0
    for word in tokenize_text:
        x = re.compile('[es|ed]$')
        if x.match(word.lower()):
            count += 0
        else:
            for char in word:
                if char.lower() in string.ascii_lowercase:
                    count += 1
        if count > 2:
            complex_word_count += 1
            count = 0
    percentage_of_complex_words = complex_word_count / len(tokenize_text)
    
    # fog Index
    fog_index = 0.4 * (avg_length + percentage_of_complex_words)
    
    # average number of words per sentence
    avg_no_of_words_per_sentence = sum(len(word.split()) for word in text.split('.')) / len(text.split('.'))
    
    # complex word count
    complex_word_count = sum(1 for word in tokenize_text if len(word) > 2)
    
    # word count
    word_count = len(tokenize_text)
    
    # syllable per word
    count = sum(1 for word in tokenize_text for char in word if char.lower() in 'aeiou')
    syllable_per_word = count / word_count
        
    # personal pronouns count
    pronouns = ['i', 'me', 'my', 'mine', 'myself', 'we', 'us', 'our', 'ours', 'ourselves']
    count = 0
    for word in text:
        if word.lower() in pronouns:
            count += 1
    personal_pronouns = count
    
    # average word length
    avg_word_length = sum(len(word) for word in tokenize_text) / word_count
    
    return positive_score, negative_score, polarity_score, subjectivity_score, avg_length, percentage_of_complex_words, fog_index, avg_no_of_words_per_sentence, complex_word_count, word_count, syllable_per_word, personal_pronouns, avg_word_length

# Function to extract and save article
def extract_and_save_article(url, url_id, folder_path, output_ws):
#     title, text = extract_text(url)
    text = extract_text(url)
    file_name = f"{url_id}.txt"
    file_path = os.path.join(folder_path, file_name)
    with open(file_path, 'w', encoding='utf-8') as file:
#         file.write(f"Title: {title}\n\n")
        file.write(text)
        
        # To calculate variables
        variables = compute_variables(text)        
        output_ws.append([
            url_id,
            url,
            *variables
        ])

# store the variable values in Output Data Structure Excel file 
output_file = "D:\\projects\\NLP_TextAnalysis\\Assignment\\Output Data Structure.xlsx"
output_wb = Workbook()
output_ws = output_wb.active
output_ws.append([
    "URL_ID",
    "URL",
    "POSITIVE SCORE", 
    "NEGATIVE SCORE", 
    "POLARITY SCORE", 
    "SUBJECTIVITY SCORE", 
    "AVG SENTENCE LENGTH", 
    "PERCENTAGE OF COMPLEX WORDS", 
    "FOG INDEX", 
    "AVG NUMBER OF WORDS PER SENTENCE", 
    "COMPLEX WORD COUNT", 
    "WORD COUNT", 
    "SYLLABLE PER WORD", 
    "PERSONAL PRONOUNS", 
    "AVG WORD LENGTH"
])

# Load input file from your desktop
input_file = "D:\\projects\\NLP_TextAnalysis\\Assignment\\Input.xlsx"
wb = load_workbook(input_file)
ws = wb.active

#Give the path to save all text files to your local desktop
folder_path = "D:\\projects\\NLP_TextAnalysis\\Assignment\\output"

for row in ws.iter_rows(min_row=2, values_only=True):
    url_id, url = row
    extract_and_save_article(url, url_id, folder_path, output_ws)

# save the variable values in the output Excel file
output_wb.save(output_file)

print("Data extraction and Excel file creation completed!")

